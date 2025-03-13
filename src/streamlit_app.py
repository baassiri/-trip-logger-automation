import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from invoice_automation import force_update_trip_log
import sys
import os

# Ensure `src/` folder is in the Python path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from config import FILE_PATH

# Ensure correct working directory
os.chdir(os.path.dirname(FILE_PATH))

# Set Streamlit page config
st.set_page_config(page_title="Trip Logger", layout="wide")

# Title
st.title("🚛 Trip Logger Automation")

# Debugging: Show current working directory
st.write(f"📂 **Current Working Directory:** `{os.getcwd()}`")
st.write(f"📄 **Excel File Path:** `{FILE_PATH}`")

if not os.path.exists(FILE_PATH):
    st.error("❌ Excel file NOT found!")
else:
    st.success("✅ Excel file found!")

# Input Fields
client_name = st.text_input("Enter Client Name")

# Dynamic address input
addresses = []
address_count = st.number_input("Number of Addresses", min_value=1, max_value=5, value=1, step=1)

for i in range(address_count):
    col1, col2 = st.columns(2)
    with col1:
        address1 = st.text_input(f"Address Line 1 - {i + 1}")
        address2 = st.text_input(f"Address Line 2 (Optional) - {i + 1}")
    with col2:
        city = st.text_input(f"City - {i + 1}")
        state = st.text_input(f"State (2-letter code) - {i + 1}")
        zip_code = st.text_input(f"ZIP Code - {i + 1}")

    if address1 and city and state and zip_code:
        full_address = f"{address1}, {address2 + ', ' if address2 else ''}{city}, {state} {zip_code}"
        addresses.append(full_address)

# Button to log the trip
if st.button("Log Trip"):
    if client_name and addresses:
        detected_clients = [client_name]
        detected_addresses = {client_name: addresses}  # Store addresses in dictionary
        force_update_trip_log(detected_clients, detected_addresses)
        st.success(f"✅ Trip logged for {client_name}")
    else:
        st.warning("⚠️ Please enter both Client Name and at least one complete Address.")

# Load and display the trip logs
st.subheader("📋 Current Trip Logs")

try:
    wb = load_workbook(FILE_PATH, data_only=True)
    ws = wb["TRIP LOGS"]

    # Extract all rows from row 7 onwards (assuming max 10 useful columns)
    data = []
    for row in ws.iter_rows(min_row=7, max_col=10, values_only=True):
        if any(row):  # Avoid empty rows
            data.append(row)

    if data:
        columns = [
            "Date", "Client", "Base", "Home Address",
            "Destination 1", "Destination 2", "Destination 3", "Destination 4", "Destination 5",
            "Extras"
        ]
        df = pd.DataFrame(data, columns=columns)

        st.dataframe(df)

        # Create a CSV export
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False)
        st.download_button("📥 Download as CSV", csv_buffer.getvalue(), "trip_logs.csv", "text/csv")
    else:
        st.info("ℹ️ No trip logs available yet.")

except Exception as e:
    st.error(f"⚠️ Could not load trip logs: {e}")
