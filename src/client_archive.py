import os
from openpyxl import load_workbook
from config import FILE_PATH

def move_to_archive():
    """
    Move clients marked as 'archive' from CLIENT DATA to CLIENT ARCHIVE.
    VBA Equivalent: MoveToArchive()
    """
    if not os.path.exists(FILE_PATH):
        print(f"❌ Error: Excel file not found at {FILE_PATH}")
        return

    try:
        wb = load_workbook(FILE_PATH, keep_vba=True)
    except Exception as e:
        print(f"❌ Error loading workbook: {e}")
        return

    try:
        ws_client_data = wb["CLIENT DATA"]
        ws_client_archive = wb["CLIENT ARCHIVE"]
    except KeyError as e:
        print(f"❌ Sheet not found: {e}")
        wb.close()
        return

    # Find last row in CLIENT DATA (Column A)
    last_row_data = ws_client_data.max_row

    # Loop backward from last row down to row 2
    for i in range(last_row_data, 1, -1):
        cell_value = ws_client_data.cell(row=i, column=21).value  # Column U
        if cell_value and str(cell_value).strip().lower() == "archive":
            # Determine next available row in CLIENT ARCHIVE
            last_row_archive = ws_client_archive.max_row + 1
            if last_row_archive < 6:
                last_row_archive = 6

            # Copy row from CLIENT DATA (columns 1 to 21)
            row_data = [ws_client_data.cell(row=i, column=col).value for col in range(1, 22)]
            
            # Paste into CLIENT ARCHIVE
            for col in range(1, 22):
                ws_client_archive.cell(row=last_row_archive, column=col, value=row_data[col - 1])

            # Verify successful copy, then delete the row if verified
            if ws_client_archive.cell(row=last_row_archive, column=1).value == ws_client_data.cell(row=i, column=1).value:
                ws_client_data.delete_rows(i)
            else:
                print(f"⚠️ Error: Row {i} could not be copied to CLIENT ARCHIVE. Deletion skipped.")

    try:
        wb.save(FILE_PATH)
        print("✅ Clients marked as 'archive' have been moved successfully!")
    except Exception as e:
        print(f"⚠️ Error saving workbook: {e}")
    finally:
        wb.close()

def restore_from_archive():
    """
    Move clients marked as 'active' from CLIENT ARCHIVE back to CLIENT DATA.
    VBA Equivalent: RestoreFromArchive()
    """
    if not os.path.exists(FILE_PATH):
        print(f"❌ Error: Excel file not found at {FILE_PATH}")
        return

    try:
        wb = load_workbook(FILE_PATH, keep_vba=True)
    except Exception as e:
        print(f"❌ Error loading workbook: {e}")
        return

    try:
        ws_client_data = wb["CLIENT DATA"]
        ws_client_archive = wb["CLIENT ARCHIVE"]
    except KeyError as e:
        print(f"❌ Sheet not found: {e}")
        wb.close()
        return

    # Find last row in CLIENT ARCHIVE (Column A)
    last_row_archive = ws_client_archive.max_row

    # Loop backward from last_row_archive down to row 6
    for i in range(last_row_archive, 5, -1):
        cell_value = ws_client_archive.cell(row=i, column=21).value  # Column U
        if cell_value and str(cell_value).strip().lower() == "active":
            # Determine next available row in CLIENT DATA
            last_row_data = ws_client_data.max_row + 1
            if last_row_data < 2:
                last_row_data = 2

            # Copy row from CLIENT ARCHIVE (columns 1 to 21)
            row_data = [ws_client_archive.cell(row=i, column=col).value for col in range(1, 22)]
            
            # Paste into CLIENT DATA
            for col in range(1, 22):
                ws_client_data.cell(row=last_row_data, column=col, value=row_data[col - 1])

            # Verify and delete the row if successful
            if ws_client_data.cell(row=last_row_data, column=1).value == ws_client_archive.cell(row=i, column=1).value:
                ws_client_archive.delete_rows(i)
            else:
                print(f"⚠️ Error: Row {i} could not be copied back to CLIENT DATA. Deletion skipped.")

    try:
        wb.save(FILE_PATH)
        print("✅ Clients marked as 'active' have been restored successfully!")
    except Exception as e:
        print(f"⚠️ Error saving workbook: {e}")
    finally:
        wb.close()
