from openpyxl import load_workbook
from config import FILE_PATH

def clear_trip_logs():
    """
    Clears the trip logs in the 'TRIP LOGS' sheet of the Excel file.
    It deletes all rows starting from row 7, ensuring headers remain intact.
    """
    try:
        # Load workbook
        wb = load_workbook(FILE_PATH, keep_vba=True)

        # Ensure the 'TRIP LOGS' sheet exists
        if "TRIP LOGS" not in wb.sheetnames:
            print("❌ 'TRIP LOGS' sheet not found!")
            wb.close()
            return

        ws = wb["TRIP LOGS"]

        # Find the last row with data
        last_row = ws.max_row

        # Ensure there's data to delete
        if last_row >= 7:
            ws.delete_rows(7, last_row - 6)  # Ensure only data rows are cleared

        # Save and close
        wb.save(FILE_PATH)
        wb.close()
        print("✅ Trip logs cleared successfully.")

    except Exception as e:
        print(f"⚠️ Error clearing trip logs: {e}")

