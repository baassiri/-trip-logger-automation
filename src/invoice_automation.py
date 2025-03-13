from openpyxl import load_workbook
from datetime import datetime
import os
from config import FILE_PATH, upload_to_drive

print("✅ Script started!")

def force_update_trip_log(detected_clients, detected_addresses):
    """Append trip logs correctly after the last used row in Column A (starting from row 7)."""
    
    if not os.path.exists(FILE_PATH):
        print(f"❌ Excel file not found: {FILE_PATH}")
        return

    try:
        wb = load_workbook(FILE_PATH, keep_vba=True)
        ws = wb["TRIP LOGS"]
    except Exception as e:
        print(f"⚠️ Error loading Excel file: {e}")
        return

    current_date = datetime.now().strftime("%m/%d/%Y")
    print(f"📅 Processing trip logs for: {current_date}")

    # Determine the last row with actual data in Column A by iterating backward
    # Find the last row that contains a valid client name in Column B
    last_used_row = 7  # Start looking from row 7
    for row in range(ws.max_row, 6, -1):  # Scan backwards
        client_name = ws.cell(row=row, column=2).value  # Column B holds client names
        if client_name and str(client_name).strip():  # If it's not empty
            last_used_row = row
            break

    # The next new entry should go **right after** the last valid client row
    new_row = last_used_row + 1


    for client in detected_clients:
        addresses = detected_addresses.get(client, [])
        row_found = None

        # Check if this client on the current date already exists (only search rows 7 to last_used_row)
        for row in range(7, last_used_row + 1):
            if ws.cell(row=row, column=1).value == current_date and ws.cell(row=row, column=2).value == client:
                row_found = row
                break

        if row_found is None:
            # Insert a new entry at new_row
            ws.cell(row=new_row, column=1, value=current_date)  # Date in Column A
            ws.cell(row=new_row, column=2, value=client)           # Client in Column B

            for i, address in enumerate(addresses[:5]):  # Up to 5 destinations (Columns E to I)
                ws.cell(row=new_row, column=5 + i, value=address)

            print(f"🆕 Created new entry for {client} at row {new_row} with addresses: {addresses}")
            new_row += 1  # Update for the next new entry
        else:
            # Append addresses to the found row – in the next available destination columns (E to I)
            for address in addresses:
                for col in range(5, 10):  # Columns 5 (E) to 9 (I)
                    if ws.cell(row=row_found, column=col).value in [None, ""]:
                        ws.cell(row=row_found, column=col, value=address)
                        print(f"📌 Added {address} to {client} at row {row_found}, column {col}")
                        break

    # (Optional) Print the last few rows for verification
    print("🔍 Verifying last few rows before saving:")
    for row in range(max(7, new_row - 5), new_row):
        row_values = [ws.cell(row=row, column=col).value for col in range(1, 11)]
        print(row_values)

    try:
        wb.save(FILE_PATH)
        wb.close()
        print("✅ Local Excel file updated successfully!")
    except Exception as e:
        print(f"⚠️ Error saving Excel file: {e}")

    upload_to_drive()

# For testing purposes:
if __name__ == "__main__":
    force_update_trip_log(['Test Client'], {'Test Client': ['Test Address']})
