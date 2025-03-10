import os
from openpyxl import load_workbook
from datetime import datetime
from config import FILE_PATH

def populate_invoice():
    """Populate the INVOICE sheet with data from TRIP LOGS."""
    if not os.path.exists(FILE_PATH):
        print(f"❌ Error: Excel file not found at {FILE_PATH}")
        return

    wb = load_workbook(FILE_PATH, keep_vba=True)
    ws_invoice = wb["INVOICE"]
    ws_trip_logs = wb["TRIP LOGS"]

    client_name = ws_invoice["B4"].value.strip() if ws_invoice["B4"].value else ""
    if not client_name:
        print("⚠️ Client name is missing in the invoice.")
        return

    last_row_trips = ws_trip_logs.max_row
    current_row = 15  # Start populating from row 15
    trip_count = 0

    for i in range(8, last_row_trips + 1):
        if str(ws_trip_logs[f"B{i}"].value).strip() == client_name:
            trip_date = ws_trip_logs[f"A{i}"].value
            if not isinstance(trip_date, datetime):
                trip_date = datetime.strptime(str(trip_date), "%m/%d/%Y")

            for dest_col in range(5, 10):  # Columns E to I
                destination = ws_trip_logs.cell(row=i, column=dest_col).value
                if destination:
                    trip_count += 1
                    ws_invoice[f"A{current_row}"] = trip_date.strftime("%m-%d-%Y")
                    ws_invoice[f"B{current_row}"] = destination
                    ws_invoice[f"C{current_row}"] = "1 Unit"
                    ws_invoice[f"D{current_row}"] = "1 Unit"
                    pickup_cost = dropoff_cost = 25

                    wait_time = ws_trip_logs.cell(row=i, column=10 + (dest_col - 5)).value
                    wait_time_cost = 25 * (int(wait_time.split(" ")[0]) if wait_time else 0)
                    ws_invoice[f"E{current_row}"] = wait_time if wait_time else "0 Unit"

                    mileage = ws_trip_logs.cell(row=i, column=15 + (dest_col - 5)).value or 0
                    mileage_cost = 3 * mileage
                    ws_invoice[f"F{current_row}"] = f"{mileage:.1f} Miles"

                    unloaded = ws_trip_logs.cell(row=i, column=20 if dest_col == 5 else 21).value or 0
                    unloaded_cost = 3 * unloaded
                    ws_invoice[f"G{current_row}"] = f"{unloaded:.1f} Miles" if unloaded else "-"

                    total_cost = pickup_cost + dropoff_cost + wait_time_cost + mileage_cost + unloaded_cost
                    ws_invoice[f"C{current_row+1}"] = pickup_cost
                    ws_invoice[f"D{current_row+1}"] = dropoff_cost
                    ws_invoice[f"E{current_row+1}"] = wait_time_cost
                    ws_invoice[f"F{current_row+1}"] = mileage_cost
                    ws_invoice[f"G{current_row+1}"] = unloaded_cost
                    ws_invoice[f"H{current_row+1}"] = total_cost

                    current_row += 2

    wb.save(FILE_PATH)
    wb.close()
    
    print(f"✅ Invoice populated successfully for {trip_count} trips!" if trip_count else "⚠️ No trips found.")

